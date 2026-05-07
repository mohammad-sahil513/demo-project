"""XLSX schema extractor for named placeholders and workbook anchors."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from modules.template.schema_models import PlaceholderDef, PlaceholderLocation, TemplateSchema


def extract_xlsx_schema(template_path: Path) -> TemplateSchema:
    wb = load_workbook(filename=str(template_path), read_only=False, data_only=False)
    try:
        placeholders: list[PlaceholderDef] = []
        seen: set[str] = set()

        # Named ranges are the strongest enterprise placeholder contract for XLSX.
        for defined_name in wb.defined_names.values():
            name = str(defined_name.name or "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            dest_text = str(defined_name.attr_text or "")
            placeholders.append(
                PlaceholderDef(
                    placeholder_id=name,
                    kind="text",
                    required=True,
                    location=PlaceholderLocation(
                        part="workbook",
                        xml_path=dest_text or f"definedName:{name}",
                        context="named-range",
                    ),
                )
            )

        sheet_anchors: list[dict[str, object]] = []
        for index, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            sheet_anchors.append(
                {
                    "name": sheet_name,
                    "index": index,
                    "max_row": int(ws.max_row or 0),
                    "max_column": int(ws.max_column or 0),
                }
            )
    finally:
        wb.close()

    return TemplateSchema(
        source_format="xlsx",
        placeholders=placeholders,
        locked_fidelity_anchors={"sheets": sheet_anchors},
    )

