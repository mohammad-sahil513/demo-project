"""Helpers for reading XLSX templates from storage paths.

We store template bytes under a ``.bin`` filename to keep the storage
layout uniform across formats. ``openpyxl.load_workbook`` validates the
file extension before peeking at the zip payload, so we open a binary file
stream and pass that to ``load_workbook`` — skipping the extension check
entirely.
"""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook


@contextmanager
def open_xlsx_workbook(
    path: Path,
    *,
    read_only: bool,
    data_only: bool,
) -> Iterator[object]:
    """Open an XLSX from disk for openpyxl.

    Template bytes are stored under a ``.bin`` filename; openpyxl rejects unknown
    extensions before inspecting the zip payload, so we always pass a binary stream.
    """
    fp = path.open("rb")
    try:
        wb = load_workbook(filename=fp, read_only=read_only, data_only=data_only)
        try:
            yield wb
        finally:
            wb.close()
    finally:
        fp.close()
