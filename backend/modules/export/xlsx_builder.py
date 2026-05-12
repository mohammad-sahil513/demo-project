"""Build or fill XLSX outputs.

Two flavors:

- **Inbuilt path** — start from an empty workbook and create one sheet
  per section. Section content (markdown tables) is parsed into native
  Excel cells with auto-fit column widths.
- **Custom path** — open the user-supplied template and append rows
  beneath the existing header layout. The sheet name matching is
  case- and length-tolerant because Excel caps sheet names at 31
  characters.

For both paths, section bodies are split into markdown blocks and each
GFM table becomes a contiguous range of cells with bold header rows.
Diagram sections are ignored — Excel exports are tabular-only.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from modules.assembly.models import AssembledDocument
from modules.export.markdown_tables import parse_gfm_table, split_markdown_blocks
from modules.template.xlsx_workbook import open_xlsx_workbook


def _sheet_title(title: str) -> str:
    t = title.strip()[:31]
    return t or "Sheet"


def _get_or_create_worksheet(wb, section_title: str, sheet_map: dict[str, object] | None):
    """Resolve sheet using workbook names first, then template sheet_map entries."""
    st_full = section_title.strip()
    st_short = _sheet_title(st_full)

    def matches(sheet_name: str) -> bool:
        sn = sheet_name.strip()
        return sn == st_full or sn == st_short or sn.lower() == st_full.lower()

    for sn in wb.sheetnames:
        if matches(sn):
            return wb[sn]

    if sheet_map and isinstance(sheet_map.get("sheets"), list):
        for item in sheet_map["sheets"]:
            nm = (item.get("name") or "").strip()
            if not nm:
                continue
            if nm == st_full or nm.lower() == st_full.lower() or nm == st_short:
                if nm in wb.sheetnames:
                    return wb[nm]
                n2 = _sheet_title(nm)
                if n2 in wb.sheetnames:
                    return wb[n2]

    new_name = st_short
    if new_name not in wb.sheetnames:
        wb.create_sheet(title=new_name)
    return wb[new_name]


def _normalized_header(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


class XlsxBuilder:
    def build(
        self,
        assembled: AssembledDocument,
        output_path: Path,
        *,
        template_path: Path | None = None,
        sheet_map: dict[str, object] | None = None,
    ) -> list[dict[str, object]]:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        warnings: list[dict[str, object]] = []
        from_template = bool(template_path and template_path.is_file())
        template_cm: object | None = None
        if from_template:
            assert template_path is not None
            template_cm = open_xlsx_workbook(template_path, read_only=False, data_only=False)
            wb = template_cm.__enter__()
        else:
            wb = Workbook()
            wb.remove(wb.active)

        try:
            for section in assembled.sections:
                ws = _get_or_create_worksheet(wb, section.title, sheet_map)
                mr = ws.max_row or 0
                mc = ws.max_column or 0
                schema_headers: list[str] = []
                schema_required: list[str] = []
                selected_header_row = 1
                if isinstance(sheet_map, dict):
                    schema_items = sheet_map.get("schema")
                    if isinstance(schema_items, list):
                        for item in schema_items:
                            if not isinstance(item, dict):
                                continue
                            if str(item.get("sheet_name") or "").strip().lower() == ws.title.strip().lower():
                                schema_headers = [str(h).strip() for h in item.get("headers", []) if str(h).strip()]
                                schema_required = [
                                    str(h).strip() for h in item.get("required_columns", []) if str(h).strip()
                                ]
                                meta = item.get("header_detection_metadata")
                                if isinstance(meta, dict):
                                    try:
                                        selected_header_row = max(1, int(meta.get("selected_row") or 1))
                                    except (TypeError, ValueError):
                                        selected_header_row = 1
                                break

                if from_template:
                    header_nonempty = mr >= selected_header_row and any(
                        ws.cell(row=selected_header_row, column=c).value not in (None, "")
                        for c in range(1, mc + 1)
                    )
                    if not header_nonempty and selected_header_row != 1:
                        header_nonempty = mr >= 1 and any(
                            ws.cell(row=1, column=c).value not in (None, "")
                            for c in range(1, mc + 1)
                        )
                        selected_header_row = 1
                        warnings.append(
                            {
                                "code": "schema_mismatch",
                                "severity": "error",
                                "section_id": section.section_id,
                                "sheet": ws.title,
                                "message": "Selected header row was empty; fell back to row 1.",
                            }
                        )
                    if mr > selected_header_row:
                        ws.delete_rows(selected_header_row + 1, mr - selected_header_row)
                    elif mr == selected_header_row and not header_nonempty:
                        ws.delete_rows(selected_header_row, 1)
                    start_row = selected_header_row + 1 if header_nonempty else selected_header_row
                else:
                    if mr > 0:
                        ws.delete_rows(1, mr)
                    start_row = 1

                rows: list[list[str]] = []
                if section.output_type == "table":
                    for kind, payload in split_markdown_blocks(section.content):
                        if kind == "table":
                            rows.extend(parse_gfm_table(payload))
                    if rows:
                        generated_headers = [str(v).strip() for v in rows[0]]
                        data_rows = rows[1:] if len(rows) > 1 else []
                        if schema_headers:
                            schema_idx = {_normalized_header(h): idx for idx, h in enumerate(schema_headers)}
                            gen_idx = {_normalized_header(h): idx for idx, h in enumerate(generated_headers)}
                            unmapped_generated = [
                                h for h in generated_headers if _normalized_header(h) not in schema_idx
                            ]
                            missing_required = [
                                h for h in (schema_required or schema_headers) if _normalized_header(h) not in gen_idx
                            ]
                            if unmapped_generated:
                                warnings.append(
                                    {
                                        "code": "unmapped_generated_columns",
                                        "severity": "warning",
                                        "section_id": section.section_id,
                                        "sheet": ws.title,
                                        "columns": unmapped_generated,
                                    }
                                )
                            if missing_required:
                                warnings.append(
                                    {
                                        "code": "missing_required_columns",
                                        "severity": "error",
                                        "section_id": section.section_id,
                                        "sheet": ws.title,
                                        "columns": missing_required,
                                    }
                                )

                            mapped_rows: list[list[str]] = []
                            mapped_rows.append(list(schema_headers))
                            for row in data_rows:
                                new_row = ["TBD"] * len(schema_headers)
                                for sh, s_idx in schema_idx.items():
                                    if sh in gen_idx and gen_idx[sh] < len(row):
                                        value = str(row[gen_idx[sh]])
                                        new_row[s_idx] = value if value else "TBD"
                                mapped_rows.append(new_row)
                            rows = mapped_rows
                        elif from_template and generated_headers:
                            warnings.append(
                                {
                                    "code": "schema_mismatch",
                                    "severity": "error",
                                    "section_id": section.section_id,
                                    "sheet": ws.title,
                                    "message": "Generated table headers could not be validated against template schema.",
                                }
                            )
                if not rows and section.content.strip():
                    rows = [[section.content.strip()]]

                for r_off, row in enumerate(rows):
                    r_idx = start_row + r_off
                    for c_idx, val in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=val)
                        if r_off == 0:
                            ws.column_dimensions[get_column_letter(c_idx)].width = min(
                                max(len(str(val)) + 2, 12),
                                48,
                            )

            wb.save(str(output_path))
            return warnings
        finally:
            if template_cm is not None:
                template_cm.__exit__(None, None, None)
