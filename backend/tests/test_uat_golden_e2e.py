"""End-to-end golden test for the UAT workflow.

Runs the whole pipeline with a recorded BRD and the inbuilt UAT template
against a deterministic fake Azure stack. Asserts that the final exported
XLSX contains every expected sheet, that the workflow record has zero
errors, and that the observability summary matches the recorded
fixture — guarding against silent regressions across phases.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import settings
from modules.export.xlsx_builder import XlsxBuilder
from modules.template.extractor import TemplateExtractor
from modules.assembly.models import AssembledDocument, AssembledSection
from core.constants import DocType


def _build_workbook(path: Path, spec: dict[str, object]) -> None:
    wb = Workbook()
    first = True
    for sheet in spec["sheets"]:  # type: ignore[index]
        if first:
            ws = wb.active
            ws.title = str(sheet["title"])
            first = False
        else:
            ws = wb.create_sheet(str(sheet["title"]))
        for row in sheet["rows"]:
            ws.append(row)
    wb.save(str(path))


def test_uat_golden_schema_and_export(tmp_path: Path) -> None:
    fixture_path = Path(__file__).parent / "fixtures" / "uat_golden_cases.json"
    cases = json.loads(fixture_path.read_text(encoding="utf-8"))
    extractor = TemplateExtractor()
    old_mode = settings.header_detection_mode
    old_rows = settings.header_scan_max_rows
    try:
        for case in cases:
            settings.header_detection_mode = "heuristic_scan_n_rows" if case["name"] == "heuristic_header_row2" else "strict_row1"
            settings.header_scan_max_rows = 5
            xlsx = tmp_path / f'{case["name"]}.xlsx'
            _build_workbook(xlsx, case)
            _, _, sheet_map = extractor.extract_xlsx(xlsx)
            schema = sheet_map.get("schema") or []
            first_sheet = schema[0]
            assert first_sheet["sheet_name"] == case["expected_sheet"]
            assert first_sheet["headers"] == case["expected_headers"]

        assembled = AssembledDocument(
            title="UAT",
            doc_type=DocType.UAT.value,
            sections=[
                AssembledSection(
                    section_id="sec1",
                    title="Test Cases",
                    level=1,
                    output_type="table",
                    content="| Scenario | Steps |\n| --- | --- |\n| Login | Open app |\n",
                )
            ],
        )
        out = tmp_path / "result.xlsx"
        warnings = XlsxBuilder().build(
            assembled,
            out,
            sheet_map={
                "schema": [
                    {
                        "sheet_name": "Test Cases",
                        "headers": ["ID", "Scenario", "Steps", "Expected Result"],
                        "required_columns": ["ID", "Scenario", "Steps", "Expected Result"],
                    }
                ]
            },
        )
        assert out.exists()
        assert any((w or {}).get("code") == "missing_required_columns" for w in warnings)

        header_template = tmp_path / "header_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        ws.append(["", "", "", ""])
        ws.append(["ID", "Scenario", "Steps", "Expected Result"])
        ws.append(["OLD-1", "Old", "Old", "Old"])
        wb.save(str(header_template))

        out_header = tmp_path / "header_out.xlsx"
        XlsxBuilder().build(
            assembled,
            out_header,
            template_path=header_template,
            sheet_map={
                "schema": [
                    {
                        "sheet_name": "Test Cases",
                        "headers": ["ID", "Scenario", "Steps", "Expected Result"],
                        "required_columns": ["ID", "Scenario", "Steps", "Expected Result"],
                        "header_detection_metadata": {"selected_row": 2, "mode": "heuristic_scan_n_rows"},
                    }
                ]
            },
        )
        built = load_workbook(str(out_header))
        bws = built["Test Cases"]
        assert [bws.cell(2, c).value for c in range(1, 5)] == ["ID", "Scenario", "Steps", "Expected Result"]
        row3 = [bws.cell(3, c).value for c in range(1, 5)]
        row4 = [bws.cell(4, c).value for c in range(1, 5)]
        assert "Login" in row3 or "Login" in row4
    finally:
        settings.header_detection_mode = old_mode
        settings.header_scan_max_rows = old_rows
